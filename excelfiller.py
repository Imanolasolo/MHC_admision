import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import sqlite3
import os

st.set_page_config(
    page_title="Formulario 001 ADMISION",
    page_icon="🩺",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://wa.me/5930993513082?text=Solicito%20ayuda%20con%20la%20app%20MHC',
        'Report a bug': "https://wa.me/5930993513082?text=Solicito%20ayuda%20con%20la%20app%20MHC",
        'About': "# App creada por CodeCodix"
    }
)

# Función para conectar a la base de datos y obtener los pacientes
def get_patients():
    conn = sqlite3.connect('hospital.db')
    query = "SELECT * FROM pacientes"
    df = pd.read_sql(query, conn)
    conn.close()
    return df

# Función para crear un nuevo Excel con los datos del paciente para ambas hojas
def create_patient_excel(template_file, patient_data):
    workbook = load_workbook(template_file)

    # Obtener ambas hojas
    if '1' in workbook.sheetnames and '2' in workbook.sheetnames:
        sheet1 = workbook['1']
        sheet2 = workbook['2']
    else:
        st.error("Una o ambas hojas '1' y '2' no existen en la plantilla de Excel.")
        return None

    # Función para establecer valor en las celdas, manejando celdas combinadas
    def set_value(sheet, cell, value):
        if pd.isna(value):
            value = ""
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                min_col, min_row, max_col, max_row = merged_range.bounds
                sheet.cell(row=min_row, column=min_col).value = value
                return
        cell.value = value

    # Llenar los datos para la hoja 1
    set_value(sheet1, sheet1['AD11'], patient_data['primer_nombre'])
    set_value(sheet1, sheet1['AQ11'], patient_data['segundo_nombre'])
    set_value(sheet1, sheet1['A11'], patient_data['primer_apellido'])
    set_value(sheet1, sheet1['P11'], patient_data['segundo_apellido'])
    set_value(sheet1, sheet1['BD11'], patient_data['tipo_documento_identificacion'])
    set_value(sheet1, sheet1['A7'], patient_data['fecha_admision'])
    set_value(sheet1, sheet1['AD7'], patient_data['sello_y_firma_del_responsable'])
    set_value(sheet1, sheet1['A13'], patient_data['estado_civil'])
    set_value(sheet1, sheet1['I13'], patient_data['sexo'])
    set_value(sheet1, sheet1['I13'], patient_data['telefono_fijo'])
    set_value(sheet1, sheet1['AD13'], patient_data['telefono_celular'])
    set_value(sheet1, sheet1['AU13'], patient_data['correo_electronico'])
    set_value(sheet1, sheet1['G18'], patient_data['provincia'])
    set_value(sheet1, sheet1['U18'], patient_data['canton'])
    set_value(sheet1, sheet1['AH18'], patient_data['parroquia'])
    set_value(sheet1, sheet1['AU18'], patient_data['barrio_sector'])
    set_value(sheet1, sheet1['G20'], patient_data['calle_principal'])
    set_value(sheet1, sheet1['Y20'], patient_data['calle_secundaria'])
    set_value(sheet1, sheet1['AS20'], patient_data['referencia'])
    set_value(sheet1, sheet1['A22'], patient_data['autocertificacion_etnica'])
    set_value(sheet1, sheet1['O22'], patient_data['nacionalidad_etnica'])
    set_value(sheet1, sheet1['AH22'], patient_data['pueblos'])
    set_value(sheet1, sheet1['AX22'], patient_data['nivel_educacion'])
    set_value(sheet1, sheet1['A24'], patient_data['estado_nivel_educacion'])
    set_value(sheet1, sheet1['P24'], patient_data['ocupacion'])
    set_value(sheet1, sheet1['AF24'], patient_data['tipo_empresa_trabajo'])
    set_value(sheet1, sheet1['AQ24'], patient_data['seguro_salud_principal'])
    set_value(sheet1, sheet1['BA24'], patient_data['tipo_bono_recibe'])
    set_value(sheet1, sheet1['A29'], patient_data['en_caso_necesario_llamar_a'])
    set_value(sheet1, sheet1['V29'], patient_data['parentesco'])
    set_value(sheet1, sheet1['AE29'], patient_data['direccion'])
    set_value(sheet1, sheet1['BB29'], patient_data['telefono_contacto'])

    # Llenar los datos para la hoja 2
    set_value(sheet2, sheet2['C6'], patient_data['fecha_admision'])
    set_value(sheet2, sheet2['C7'], patient_data['fecha_egreso'])
    set_value(sheet2, sheet2['I6'], patient_data['numero_dis_estadia'])
    set_value(sheet2, sheet2['L6'], patient_data['codigo_de_servicio_INEC'])
    set_value(sheet2, sheet2['Q6'], patient_data['diagnosticos_o_sindromes_principal'])
    set_value(sheet2, sheet2['AF6'], patient_data['CIE'])
    set_value(sheet2, sheet2['AJ6'], patient_data['definitivo'])
    set_value(sheet2, sheet2['AL6'], patient_data['diagnosticos_o_sindromes_secundario'])
    set_value(sheet2, sheet2['BA6'], patient_data['CIE_secundario'])
    set_value(sheet2, sheet2['BE6'], patient_data['definitivo_secundario'])
    set_value(sheet2, sheet2['BG6'], patient_data['alta'])
    set_value(sheet2, sheet2['BI6'], patient_data['muerte_menos_de_48_horas'])
    set_value(sheet2, sheet2['BK6'], patient_data['muerte_mas_de_48_horas'])
    set_value(sheet2, sheet2['BM6'], patient_data['clinico'])
    set_value(sheet2, sheet2['BQ6'], patient_data['quirurgico'])
    set_value(sheet2, sheet2['CE6'], patient_data['sello_y_firma_del_responsable'])

    # Guardar el nuevo archivo Excel en memoria
    excel_file_path = "paciente.xlsx"
    workbook.save(excel_file_path)

    return excel_file_path

# Interfaz de Streamlit
st.title("Generador de Excel de Pacientes")

# Cargar los pacientes desde la base de datos
patients_df = get_patients()

# Seleccionar paciente
patient_id = st.selectbox("Selecciona un paciente", patients_df['id'].tolist(), 
    format_func=lambda x: f"{patients_df[patients_df['id'] == x]['primer_nombre'].values[0]} {patients_df[patients_df['id'] == x]['primer_apellido'].values[0]}")

# Botón para generar el archivo Excel
if st.button("Generar Excel"):
    selected_patient = patients_df[patients_df['id'] == patient_id].iloc[0]
    
    # Depuración: Mostrar los datos del paciente seleccionado
    st.write("Datos del paciente seleccionado:")
    st.write(selected_patient)
    
    # Cargar la plantilla de Excel
    template_file = "admision.xlsx"
    
    # Crear el nuevo archivo Excel con los datos del paciente para ambas hojas
    new_file_name = create_patient_excel(template_file, selected_patient)

    # Mensaje de éxito y botón de descarga
    if new_file_name and os.path.exists(new_file_name):
        st.success("El archivo Excel se ha creado correctamente.")
        with open(new_file_name, "rb") as file:
            st.download_button("Descargar el archivo generado", file, new_file_name)



